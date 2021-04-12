class County_Zipcodes:

	import zipcodes #pip install zipcodes
	from openpyxl import load_workbook #pip install openpyxl

	@staticmethod
	def generate_zipcodes(FILENAME, COUNTY, OUTPUT=None, WORKSHEET=None, WORKSHEET_COLUMN=None):
		import zipcodes
		from openpyxl import load_workbook

		if WORKSHEET is None:
			file_Worksheet = "Sheet1"
		else: file_Worksheet = WORKSHEET

		if WORKSHEET_COLUMN is None:
			file_worksheet_Column = 'A'
		else: file_worksheet_Column = WORKSHEET_COLUMN

		if OUTPUT is None:
			file_Output = str(COUNTY)+" Zipcodes.txt"
		else: file_Output = OUTPUT

		temp_Dict = {}

		try:
			file_Load = load_workbook(FILENAME)
		except:
			print("Error on File Load")
			return

		try:
			file_Worksheet = file_Load.get_sheet_by_name(file_Worksheet)
		except:
			print("Error on File Worksheet name")
			return


		file_Load_Column = file_Worksheet[file_worksheet_Column]

		file_Loaded_List = [file_Load_Column[x].value for x in range(len(file_Load_Column))]

		# O(N) notation currently
		# will try and add to the rework the Zipcodes import to add functionality for a bool matching function
		# for o(1) notation
		# Credit: https://github.com/seanpianka/Zipcodes

		for x in file_Loaded_List:

			y = zipcodes.similar_to(x, zips=zipcodes.filter_by(active=True, county=COUNTY))

			if len(y) > 0:
				if x in temp_Dict:
					temp_Dict[x] += 1
				else:
					temp_Dict[x] = 1

		County_Zipcodes.saveToFile(file_Output, COUNTY, temp_Dict)


	def saveToFile(OUTPUT, County, Dictionary):

		with open(OUTPUT, 'w') as f:

			f.writelines(f'These are the zipcodes valid in the {County} area.\n')

			for k,v in Dictionary.items():
				f.writelines(f'{k}: {v}\n')

			f.close()

